"""XLSX workbook content extractor using openpyxl in low-memory read_only streaming mode."""

from pathlib import Path
from typing import Generator, Union

import openpyxl

from src.ingestion.detectors.base import DetectionResult, FileType
from src.ingestion.extractors.base import BaseExtractor
from src.ingestion.extractors.models import ContentType, ExtractedElement, ExtractedStream, ExtractionStatus


class XLSXExtractor(BaseExtractor):
    """Extracts Excel worksheet rows in streaming mode while preserving formula strings."""

    def __init__(self):
        super().__init__(target_file_type=FileType.XLSX)

    def _extract_implementation(self, path: Path, detection_result: DetectionResult) -> ExtractedStream:
        try:
            wb = openpyxl.load_workbook(path, read_only=True, data_only=False)
            metadata = {
                "sheet_names": wb.sheetnames,
                "formula_preservation_policy": "data_only=False (raw formulas preserved)",
            }

            generator = self._row_generator(wb)
            return ExtractedStream(
                file_path=str(path),
                file_type=FileType.XLSX,
                element_generator=generator,
                metadata=metadata,
            )

        except Exception as e:
            return ExtractedStream(
                file_path=str(path),
                file_type=FileType.XLSX,
                element_generator=iter([]),
                status=ExtractionStatus.FAILED,
                error_message=f"Failed to open XLSX workbook: {str(e)}",
            )

    def _row_generator(self, wb: openpyxl.Workbook) -> Generator[ExtractedElement, None, None]:
        element_idx = 0

        try:
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                for r_idx, row in enumerate(ws.iter_rows(values_only=False), start=1):
                    row_values = []
                    has_content = False

                    for cell in row:
                        val = cell.value
                        if val is not None:
                            has_content = True
                            row_values.append(str(val).strip())

                    if not has_content:
                        continue

                    row_text = " | ".join(row_values)
                    yield ExtractedElement(
                        element_index=element_idx,
                        content_type=ContentType.SHEET_ROW,
                        text_content=row_text,
                        structural_context={
                            "sheet_name": sheet_name,
                            "row_index": r_idx,
                            "col_count": len(row_values),
                            "cached_evaluated_value_guaranteed": False,
                        },
                    )
                    element_idx += 1
        finally:
            wb.close()
